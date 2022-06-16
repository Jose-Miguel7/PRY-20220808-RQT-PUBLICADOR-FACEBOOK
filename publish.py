import time
import random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import sys
import json
from openpyxl import load_workbook


def sleep_time():
    time.sleep(random.uniform(5, 10))


class BotFacebookMarketplace:
    def __init__(self, excel, email, password):
        self.excel = excel
        self.email = email
        self.password = password

        opts = Options()
        opts.add_argument(
            "user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Ubuntu Chromium/71.0.3578.80 Chrome/71.0.3578.80 Safari/537.36")

        if sys.platform == "win32":
            self.driver = webdriver.Chrome('./chromedriver.exe', options=opts)
        else:
            self.driver = webdriver.Chrome('./chromedriver', options=opts)
        self.driver.maximize_window()

        self.get_access_facebook()
        sleep_time()
        self.iterate_excel()

    def change_page(self, url):
        self.driver.get(url)
        sleep_time()

    def get_access_facebook(self):
        self.change_page("https://www.facebook.com")
        email_input = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "email")))
        email_input.send_keys(self.email)
        sleep_time()
        password_input = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "pass")))
        password_input.send_keys(self.password)
        sleep_time()
        login_button = WebDriverWait(self.driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@type='submit']")))
        login_button.click()

    def select_category(self, category):
        with open('./data/category.json', encoding="utf8") as r:
            data = json.loads(r.read())
            data_category = data[int(category)]
            category_label = data_category['category']

        try:
            category_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@aria-label='Categoría']")))
            category_input.click()
            sleep_time()
            category_option = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
                (By.XPATH, f"//*[text() ='{category_label}']/ancestor::div[@role='button']")))
            category_option.click()
            sleep_time()
        except TimeoutException:
            category_input = WebDriverWait(self.driver, random.uniform(7, 15)).until(EC.presence_of_element_located(
                (By.XPATH, "//label[@aria-label='Categoría']")))
            category_input.click()
            sleep_time()
            category_option = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
                (By.XPATH, f"//*[text() ='{category_label}']/ancestor::div[@role='button']")))
            category_option.click()
            sleep_time()

    def send_text_data(self, title, price, description, sku):
        title_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//label[@aria-label='Título']/div/div/input")))
        title_input.send_keys(title)

        sleep_time()

        price_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//label[@aria-label='Precio']/div/div/input")))
        price_input.send_keys(price)

        sleep_time()

        description_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//label[@aria-label='Descripción']/div/div/textarea")))
        description_input.send_keys(description.replace("\r\n", "\n"))

        sleep_time()

        sku_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//label[@aria-label='SKU']/div/div/input")))
        sku_input.send_keys(sku)

        sleep_time()

    def select_state(self):
        state_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//label[@aria-label='Estado']")))
        state_input.click()
        sleep_time()
        state_option = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, f"//*[text() ='Nuevo']/ancestor::div[@role='option']")))
        state_option.click()
        sleep_time()

    def select_ubication(self, ubication):
        ubication_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//input[@aria-label='Ingresa una ciudad']")))
        ubication_input.clear()

        ubication_input.send_keys(ubication)

        sleep_time()

        ubication_option = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//ul[@role='listbox']/li[1]")))
        ubication_option.click()
        sleep_time()

    def upload_images(self, text_images):
        images_input = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
            (By.XPATH, "//input[@accept='image/*,image/heif,image/heic']")))
        images_input.send_keys(text_images)
        sleep_time()

    def iterate_excel(self):
        wb = load_workbook(self.excel)
        ws = wb.active

        counter = 0
        success = 0

        for rows in ws.iter_rows(min_row=1, min_col=2):
            counter += 1

            try:
                sku = rows[0].value if rows[0].value else ''
                category = str(rows[1].value)
                title = rows[2].value
                description = rows[3].value
                price_detail = str(rows[4].value)
                image1 = rows[5].value
                image2 = rows[6].value
                image3 = rows[7].value
                image4 = rows[8].value
                image5 = rows[9].value
                image6 = rows[10].value
                region_detail = str(rows[11].value)
                comuna_detail = str(rows[12].value)
                images = (image2, image3, image4, image5, image6)
            except Exception as e:
                print(e)
                sku, category, title, description, price_detail, image1, image2, image3 = None, None, None, None, None, None, None, None
                image4, image5, image6, region_detail, comuna_detail, images = None, None, None, None, None, None
            if image1:
                text_images = image1
                for image in images:
                    if image:
                        text_images = text_images + ' \n ' + image

            else:
                text_images = ''

            if category and title and description and price_detail and region_detail and comuna_detail:
                self.change_page("https://www.facebook.com/marketplace/create/item")

                # ----------------------------  SELECT CATEGORY  --------------------------------
                self.select_category(category)

                # ----------------------------  TITLE - DESCRIPTION - PRICE  ----------------------------
                self.send_text_data(title, price_detail, description, sku)

                # ----------------------------  REGION - COMUNA  ----------------------------
                self.select_ubication(f"{region_detail}, {comuna_detail}")

                # ---------------------------- IMAGES  ----------------------------
                self.upload_images(text_images)
                sleep_time()

                self.select_state()

                # ----------------------------  NEXT BUTTON  ----------------------------
                next_button = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.presence_of_element_located(
                    (By.XPATH, "//div[@aria-label='Siguiente']")))
                next_button.click()

                sleep_time()

                post_button = WebDriverWait(self.driver, random.uniform(8, 15)).until(EC.element_to_be_clickable(
                    (By.XPATH, "//div[@aria-label='Publicar']")))
                post_button.click()
                sleep_time()

        self.driver.quit()
        return counter, success
